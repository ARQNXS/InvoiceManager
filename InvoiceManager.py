import os
from datetime import datetime, timedelta
import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import Canvas
from openpyxl import load_workbook
import pdfkit


class InvoiceManager:
    def __init__(self, template_path, booking_file):
        self.template_path = template_path
        self.booking_file = booking_file
        
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file {template_path} not found.")
        
        self.wb_template = openpyxl.load_workbook(template_path, keep_links=False)
        self.template_sheet = self.wb_template.active
        
        if not os.path.exists(booking_file):
            self.df_booking = pd.DataFrame(columns=['Invoice Number', 'Name', 'Amount', 'Date', 'Due Date', 'File Path',
                                                    'Address', 'City', 'Postal Code', 'Country', 'Phone Number',
                                                    'Hourly Rate', 'Hours Booked', 'Status'])
            self.df_booking.to_csv(booking_file, index=False)
        else:
            self.df_booking = pd.read_csv(booking_file)
            
            if 'Invoice Number' not in self.df_booking.columns or self.df_booking['Invoice Number'].empty:
                self.df_booking = pd.DataFrame(columns=['Invoice Number', 'Name', 'Amount', 'Date', 'Due Date', 'File Path',
                                                        'Address', 'City', 'Postal Code', 'Country', 'Phone Number',
                                                        'Hourly Rate', 'Hours Booked', 'Status'])
                self.df_booking.to_csv(booking_file, index=False)

        if self.df_booking['Invoice Number'].empty or self.df_booking['Invoice Number'].isna().all():
            self.next_invoice_number = 1
        else:
            self.next_invoice_number = self.df_booking['Invoice Number'].apply(lambda x: int(x[1:]) if x.startswith('s') else 0).max() + 1

    

    def create_invoice(self, name, amount, date, due_date=None, hours=None, hourly_rate=None, total=None,
                       address=None, city=None, postal_code=None, country=None, phone_number=None, description=None):
        try:
            # Ensure date is datetime object
            if not isinstance(date, datetime):
                date = datetime.strptime(date, "%Y-%m-%d")

            # Calculate due date if not provided
            if due_date is None:
                due_date = date + timedelta(days=14)

            # Generate invoice number and format as string
            invoice_number = max(1, self.next_invoice_number)
            invoice_number_str = f's{invoice_number}'
            print(f"Generated invoice number: {invoice_number_str} and type {type(invoice_number_str)}")

            # Update self.next_invoice_number for next invoice
            self.next_invoice_number += 1

            # Calculate amount based on provided data
            if hours is not None and hourly_rate is not None:
                amount = hours * hourly_rate
                amount = int(amount)
            elif total is not None:
                amount = total
                amount = int(amount)

            print(f"Calculated amount: {amount}")

            # Load the template
            print("Loading template...")
            wb_template = load_workbook(self.template_path)
            invoice_sheet = wb_template.active

            # Calculate due date for invoice
            due_date = date + timedelta(days=14)

            # Debugging information before setting cell values
            print("Setting invoice sheet values...")
            
            # Set values in specific cells
            invoice_sheet['B22'].value = description
            invoice_sheet['F12'].value = invoice_number_str
            invoice_sheet['C22'].value = date.strftime("%Y-%m-%d")
            invoice_sheet['B9'].value = date.strftime("%Y-%m-%d")
            invoice_sheet['D22'].value = str(date.isocalendar()[1])
            invoice_sheet['E22'].value = hours if hours is not None else ''
            invoice_sheet['F22'].value = f'€ {hourly_rate:,.2f}' if hourly_rate is not None else ''
            invoice_sheet['B12'].value = name
            invoice_sheet['B13'].value = address if address else ''
            invoice_sheet['B14'].value = city if city else ''
            invoice_sheet['B15'].value = f'{postal_code}, {country}' if postal_code and country else ''
            invoice_sheet['B16'].value = phone_number if phone_number else ''
            invoice_sheet['F17'].value = due_date.strftime("%Y-%m-%d")

           # Save the invoice to a file
            invoice_filename = f'invoice_{invoice_number_str}_{name}.xlsx'
            full_invoice_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), invoice_filename)
            print(f"Saving invoice to {full_invoice_path}...")

            try:
                wb_template.save(full_invoice_path)
                print("Invoice saved successfully.")
            except PermissionError as pe:
                print(f"Permission error: Unable to save the file. Please check if you have write access to {full_invoice_path}")
                raise pe
            except Exception as save_error:
                print(f"Error saving the file: {save_error}")
                raise save_error
            finally:
                wb_template.close()  # Ensure workbook is closed after saving
            """
            # Save as PDF in the same folder
            pdf_filename = f'invoice_{invoice_number_str}_{name}.pdf'
            full_pdf_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), pdf_filename)

             # Read Excel file into pandas DataFrame
            df = pd.read_excel(full_invoice_path)

            # Convert DataFrame to HTML
            html_filename = "temp.html"
            df.to_html(html_filename, index=False)  # Set index=False to exclude row numbers

            # Convert HTML to PDF using pdfkit
            pdfkit.from_file(html_filename, full_pdf_path)

            # Remove temporary HTML file
            os.remove(html_filename)

            print(f"Invoice PDF saved successfully to {full_pdf_path}")
            """
            # Update data frame with new invoice information
            new_invoice = pd.DataFrame([{
                'Invoice Number': invoice_number_str,
                'Name': name,
                'Description': description,
                'Amount': amount,
                'Date': date.strftime("%Y-%m-%d"),
                'Due Date': due_date.strftime("%Y-%m-%d"),
                'File Path': full_invoice_path,
                'Address': address,
                'City': city,
                'Postal Code': postal_code,
                'Country': country,
                'Phone Number': phone_number,
                'Hourly Rate': hourly_rate if hourly_rate is not None else 0,
                'Hours Booked': hours if hours is not None else 0,
                'Status': 'Outstanding'
            }])
            self.df_booking = pd.concat([self.df_booking, new_invoice], ignore_index=True)

            # Save updated data frame to CSV file
            print(f"Saving booking data to {self.booking_file}...")
            self.df_booking.to_csv(self.booking_file, index=False)
            print("Booking data saved successfully.")

            return full_invoice_path

        except Exception as e:
            print(f"Error occurred while creating invoice: {e}")
            raise e


    def send_reminders(self):
        today = datetime.today()
        due_in_7_days = today + timedelta(days=7)
        overdue = self.df_booking[pd.to_datetime(self.df_booking['Due Date']) < today]
        due_soon = self.df_booking[pd.to_datetime(self.df_booking['Due Date']) == due_in_7_days]

        reminders = []
        for _, row in overdue.iterrows():
            reminders.append(f"Reminder: Invoice {row['Invoice Number']} is overdue! Please contact {row['Name']}.")

        for _, row in due_soon.iterrows():
            reminders.append(f"Reminder: Invoice {row['Invoice Number']} is due in 7 days. Please contact {row['Name']}.")

        return reminders

    def get_invoices(self):
        return self.df_booking[['Invoice Number', 'Name', 'Amount', 'Date', 'Due Date', 'Status']]

    def set_cell_value(self, sheet, cell, value):
        try:
            print(f"Setting cell {cell} with cell type {type(cell)} with value: {value}")

            if isinstance(cell, str):
                cell = sheet[cell]
                print(f"Cell coordinates resolved: {cell.coordinate}")

            merged_range = None
            for range_coord in sheet.merged_cells.ranges:
                if cell.coordinate in range_coord:
                    merged_range = range_coord
                    print(f"Found merged range: {merged_range}")
                    break

            if merged_range:
                print(f"Unmerging cells in range: {merged_range}")
                sheet.unmerge_cells(merged_range)
                sheet[cell.coordinate] = value  # Set value directly without merging
                print(f"Set cell {cell} with value {value}")
            else:
                if isinstance(value, (int, float)):
                    print(f"Setting numeric value: {value}")
                    cell.value = value  # Set numeric value directly
                else:
                    print(f"Setting string value: {value}")
                    cell.value = str(value)  # Convert non-numeric values to string

        except Exception as e:
            print(f"Error occurred while setting cell {cell} with value {value}: {e}")
            # Optionally, raise the error again to propagate it up or handle it as needed
            raise e


    def get_total_received(self):
        return self.df_booking[self.df_booking['Status'] == 'Paid']['Amount'].sum()

    def get_total_outstanding(self):
        return self.df_booking[self.df_booking['Status'] == 'Outstanding']['Amount'].sum()

    def update_invoice_status(self, invoice_number, status):
        self.df_booking.loc[self.df_booking['Invoice Number'] == invoice_number, 'Status'] = status
        self.df_booking.to_csv(self.booking_file, index=False)

    def search_invoices(self, query):
        return self.df_booking[self.df_booking['Name'].str.contains(query, case=False, na=False)]

    def get_invoice_details(self, invoice_number):
        invoice_details = self.df_booking[self.df_booking['Invoice Number'] == invoice_number].to_string(index=False)
        return invoice_details

class InvoiceApp:
    def __init__(self, root, manager):
        self.manager = manager
        self.root = root
        self.root.title("Invoice Manager")
        self.root.geometry("1600x1000")

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        ttk.Label(main_frame, text="Name Recipient:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.name_entry = ttk.Entry(main_frame)
        self.name_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="Address:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.address_entry = ttk.Entry(main_frame)
        self.address_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="City:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.city_entry = ttk.Entry(main_frame)
        self.city_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="Postal Code:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        self.postal_code_entry = ttk.Entry(main_frame)
        self.postal_code_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="Country:").grid(row=4, column=0, sticky=tk.W, padx=5, pady=5)
        self.country_entry = ttk.Entry(main_frame)
        self.country_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="Phone Number:").grid(row=5, column=0, sticky=tk.W, padx=5, pady=5)
        self.phone_number_entry = ttk.Entry(main_frame)
        self.phone_number_entry.grid(row=5, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="Amount:").grid(row=6, column=0, sticky=tk.W, padx=5, pady=5)
        self.amount_entry = ttk.Entry(main_frame)
        self.amount_entry.grid(row=6, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="Hourly Rate:").grid(row=7, column=0, sticky=tk.W, padx=5, pady=5)
        self.hourly_rate_entry = ttk.Entry(main_frame)
        self.hourly_rate_entry.grid(row=7, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="Hours Booked:").grid(row=8, column=0, sticky=tk.W, padx=5, pady=5)
        self.hours_entry = ttk.Entry(main_frame)
        self.hours_entry.grid(row=8, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="Date (YYYY-MM-DD):").grid(row=9, column=0, sticky=tk.W, padx=5, pady=5)
        self.date_entry = ttk.Entry(main_frame)
        self.date_entry.grid(row=9, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Label(main_frame, text="Description:").grid(row=10, column=0, sticky=tk.W, padx=5, pady=5)
        self.description_entry = ttk.Entry(main_frame)
        self.description_entry.grid(row=10, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Button(main_frame, text="Create Invoice", command=self.create_invoice).grid(row=11, column=0, columnspan=2, padx=5, pady=5)

        self.invoice_list = ttk.Treeview(main_frame, columns=('Invoice Number', 'Name', 'Amount', 'Date', 'Due Date', 'Status'), show='headings')
        self.invoice_list.heading('Invoice Number', text='Invoice Number')
        self.invoice_list.heading('Name', text='Name')
        self.invoice_list.heading('Amount', text='Amount')
        self.invoice_list.heading('Date', text='Date')
        self.invoice_list.heading('Due Date', text='Due Date')
        self.invoice_list.heading('Status', text='Status')

        self.invoice_list.grid(row=12, column=0, columnspan=2, padx=5, pady=5, sticky=(tk.W, tk.E))

        self.update_invoice_list()

        ttk.Button(main_frame, text="Mark as Paid", command=self.mark_as_paid).grid(row=13, column=0, columnspan=2, padx=5, pady=5)

        ttk.Label(main_frame, text="Search:").grid(row=14, column=0, sticky=tk.W, padx=5, pady=5)
        self.search_entry = ttk.Entry(main_frame)
        self.search_entry.grid(row=14, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        ttk.Button(main_frame, text="Search", command=self.search_invoices).grid(row=14, column=2, padx=5, pady=5)

        self.detail_frame = ttk.Frame(main_frame, padding="10")
        self.detail_frame.grid(row=15, column=0, columnspan=2, padx=5, pady=5, sticky=(tk.W, tk.E))

        ttk.Label(self.detail_frame, text="Invoice Details", font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=5)
        self.invoice_detail_text = tk.Text(self.detail_frame, height=10, width=60)
        self.invoice_detail_text.grid(row=1, column=0, columnspan=2, padx=5, pady=5)

        self.invoice_list.bind("<Double-1>", self.show_invoice_details)

        self.update_invoice_list()
        self.update_visualization()

    def create_invoice(self):
        name = self.name_entry.get()
        address = self.address_entry.get()
        city = self.city_entry.get()
        postal_code = self.postal_code_entry.get()
        country = self.country_entry.get()
        phone_number = self.phone_number_entry.get()
        amount = float(self.amount_entry.get()) if self.amount_entry.get() else None
        hourly_rate = float(self.hourly_rate_entry.get()) if self.hourly_rate_entry.get() else None
        hours = float(self.hours_entry.get()) if self.hours_entry.get() else None
        date_str = self.date_entry.get()
        description = self.description_entry.get()

        if not date_str:
            messagebox.showerror("Error", "Date field cannot be empty.")
            return

        try:
            date = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Date format is incorrect. Please use YYYY-MM-DD.")
            return

        due_date = date + timedelta(days=14)  # Due date is 14 days from the invoice date

        try:
            invoice_path = self.manager.create_invoice(name, amount, date, due_date, hours, hourly_rate, amount, address, city, postal_code, country, phone_number, description)
            messagebox.showinfo("Invoice Created", f"Invoice successfully created at: {invoice_path}")
            self.update_invoice_list()
            self.clear_entries()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create invoice: {e}")

    def clear_entries(self):
        self.name_entry.delete(0, tk.END)
        self.address_entry.delete(0, tk.END)
        self.city_entry.delete(0, tk.END)
        self.postal_code_entry.delete(0, tk.END)
        self.country_entry.delete(0, tk.END)
        self.phone_number_entry.delete(0, tk.END)
        self.amount_entry.delete(0, tk.END)
        self.hourly_rate_entry.delete(0, tk.END)
        self.hours_entry.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)
        self.description_entry.delete(0, tk.END)

    def mark_as_paid(self):
        selected_item = self.invoice_list.selection()
        if not selected_item:
            return
        invoice_number = self.invoice_list.item(selected_item)['values'][0]
        self.manager.update_invoice_status(invoice_number, 'Paid')
        self.update_invoice_list()

    def update_invoice_list(self):
        for i in self.invoice_list.get_children():
            self.invoice_list.delete(i)

        invoices = self.manager.get_invoices()
        for _, row in invoices.iterrows():
            self.invoice_list.insert('', 'end', values=(row['Invoice Number'], row['Name'], row['Amount'], row['Date'],
                                                        row['Due Date'], row['Status']))

    def search_invoices(self):
        query = self.search_entry.get().strip()
        if not query:
            self.update_invoice_list()  # Reset to show all invoices if search query is empty
            return

        try:
            filtered_invoices = self.manager.search_invoices(query)
            self.invoice_list.delete(*self.invoice_list.get_children())

            for index, row in filtered_invoices.iterrows():
                self.invoice_list.insert('', 'end', values=(row['Invoice Number'], row['Name'], row['Amount'],
                                                            row['Date'], row['Due Date'], row['Status']))
        except Exception as e:
            messagebox.showerror("Search Error", f"Error occurred during search: {e}")



    def show_invoice_details(self, event):
        # Get the selected item in the invoice list
        selected_item = self.invoice_list.selection()
        
        # Check if an item is selected
        if not selected_item:
            return
        
        try:
            # Retrieve the invoice number from the selected item
            invoice_number = self.invoice_list.item(selected_item)['values'][0]
            
            # Fetch detailed information about the invoice using the invoice number
            invoice_details = self.manager.get_invoice_details(invoice_number)
            
            # Clear the current text in the invoice_detail_text widget
            self.invoice_detail_text.delete('1.0', tk.END)
            
            # Insert the fetched invoice details into the text widget
            self.invoice_detail_text.insert(tk.END, invoice_details)
        
        except Exception as e:
            # Display an error message if fetching details fails
            messagebox.showerror("Error", f"Failed to fetch invoice details: {e}")



    def update_visualization(self):
        # Create a canvas widget with specified dimensions
        canvas = Canvas(self.root, width=1200, height=400)
        canvas.grid(row=1, column=0)
        
        # Retrieve total received and total outstanding amounts from the manager
        total_received = self.manager.get_total_received()
        total_outstanding = self.manager.get_total_outstanding()
        
        # Display text on the canvas showing total received and total outstanding amounts
        canvas.create_text(100, 100, text=f"Total Received: € {total_received:,.2f}", font=("Arial", 14))
        canvas.create_text(100, 150, text=f"Total Outstanding: € {total_outstanding:,.2f}", font=("Arial", 14))



def main():
    root = tk.Tk()
    current_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(current_dir, 'template.xlsx')
    booking_file = os.path.join(current_dir, 'invoice_booking.csv')
    manager = InvoiceManager(template_path, booking_file)
    app = InvoiceApp(root, manager)
    root.mainloop()

if __name__ == "__main__":
    main()